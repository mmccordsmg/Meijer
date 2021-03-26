from meijer_funcs import clip_meijer_coupon, init_meijer_connection_user, unclip_meijer_coupon, get_all_clipped_meijer_coupons, get_meijer_offers, strip_tags
import argparse
from tqdm import trange, tqdm
import json
import random
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from zipfile import BadZipFile


parser = argparse.ArgumentParser(description="Finds  Meijer coupons for account specified")
parser.add_argument("-u",	"--username",		help="Username (Email address)",								required=True,	type=str)
parser.add_argument("-p",	"--password",		help="Password",																required=True,	type=str)
parser.add_argument("-s",	"--startnum",		help="starting number (10 digits)",							required=True,	type=str)
parser.add_argument("-e",	"--endnum",			help="ending number (10 digits)",								required=True,	type=str)
parser.add_argument("-w",	"--workbook",		help="xlsx workbook",														required=True,	type=str)
parser.add_argument("-r",	"--recheck",		help="Recheck entire range (even if results are in workbook)", required=False, action='store_true')

args = parser.parse_args()

tabs = [
						{"Name": "Valid Coupon", "Color": "009900"}, 
						{"Name": "Earned Reward", "Color": "009900"}, 
						{"Name": "InProgress Reward", "Color": "FF8000"}, 
						{"Name": "Available Reward", "Color": "FF8000"}, 
						{"Name": "Other", "Color": "FF0000"}
				]
columns = [
						{'Name': 'meijerOfferId', 'Width': 12, 'Column': 1, 'Type': 'General'},
						{'Name': 'description', 'Width': 80, 'Column': 2, 'Type': 'General'},
						{'Name': 'startDate', 'Width': 12, 'Column': 3, 'Type': 'Date'},
						{'Name': 'expirationDate', 'Width': 12, 'Column': 4, 'Type': 'Date'},
						{'Name': 'manufacturerCoupon', 'Width': 20, 'Column': 5, 'Type': 'Boolean'},
						{'Name': 'status', 'Width': 20, 'Column': 6, 'Type': 'General'},
						{'Name': 'ScanDate', 'Width': 20, 'Column': 7, 'Type': 'DateTime'},
						
				]

result_mapping = 	{
											"Valid Coupon": "Valid Coupon",
											"Earned Reward": "Earned Reward",
											"InProgress Reward": "InProgress Reward",
											"Available Reward": "Available Reward"
									}

date_style = NamedStyle(name='date', number_format='YYYY/MM/DD')
datetime_style = NamedStyle(name='datetime', number_format='YYYY/MM/DD h:mm:ss AM/PM')

def init_meijer_workbook(filename):
	wb = Workbook()
	wb.remove(wb.active)
	for tab in tabs:
		ws = wb.create_sheet(title=tab['Name'])
		ws.sheet_properties.tabColor = tab['Color']
	for ws in wb:
		for column in columns:
			c = ws.cell(row=1, column=column['Column'], value=column['Name'])
			c.font = Font(bold=True)
			ws.column_dimensions[get_column_letter(column['Column'])].width = column['Width']
		ws.auto_filter.ref = f"A:{get_column_letter(len(columns))}"
	wb.save(filename = filename)
	return(wb)

def load_meijer_workbook(filename):
	sheets = {}
	try:
		wb = load_workbook(filename = filename)
	except BadZipFile:
		print(f'Invalid workbook: {filename}')
	except FileNotFoundError:
		print(f'File {args.workbook} not found, initializing new file...')
		wb = init_meijer_workbook(filename)
	try:
		for tab in tabs:
			sheets[tab['Name']] = wb[tab['Name']]
	except KeyError as e:
		print(f'File {args.workbook} invalid: {e}')
		return(None)
	try:
		wb.add_named_style(date_style)
		wb.add_named_style(datetime_style)
	except ValueError:
		pass
	return(wb)

def load_meijer_worksheet_data(ws):
	data = []
	for value in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns), values_only=True):
		offerdata = {}
		for x in range(len(columns)):
			if (value[x] != None):
				if columns[x]['Type'] == 'Date':
					offerdata[columns[x]["Name"]] = datetime.strftime(value[x], "%Y-%m-%dT00:00:00")
				elif columns[x]['Type'] == 'DateTime':
					offerdata[columns[x]["Name"]] = datetime.strftime(value[x], "%Y-%m-%dT%H:%M:%S")
				else:
					offerdata[columns[x]["Name"]] = value[x]
		data.append(offerdata)
	return(data)

def load_meijer_workbook_data(wb):
	data = []
	for tab in tabs:
		ws = wb[tab['Name']]
		if ws.max_row > 1:
			data.extend(load_meijer_worksheet_data(ws))
	return(data)

def process_coupons(wb):
	global results
	coupons = get_all_clipped_meijer_coupons()
	if coupons:
		for coupon in coupons:
			if coupon.get("meijerOfferId") in checkrange:
				results.append({'meijerOfferId': coupon.get("meijerOfferId"), 'status': 'Valid Coupon', 'description': f'{coupon.get("title").strip()} {coupon.get("description").strip()}', 'expirationDate': coupon.get("redemptionEndDate"), 'startDate': coupon.get("redemptionStartDate"), 'manufacturerCoupon': coupon.get("manufacturerCoupon"), 'ScanDate': datetime.now().strftime('%Y-%m-%dT%H:%M:%S')})
				unclip_meijer_coupon(coupon.get("meijerOfferId"))
	return(save_results(wb))

def save_results(wb):
	global results
	clear_meijer_workbook(wb)
	for result in results:
		targetsheet = result_mapping.get(result['status'])
		if not targetsheet:
			targetsheet = 'Other'
		resultoutput = []
		for column in columns:
			if column['Type'] == 'Date' and (datefield := result.get(column['Name'])):
				formatteddata = datetime.strptime(datefield.split('T')[0], '%Y-%m-%d')
			elif column['Type'] == 'DateTime' and (datefield := result.get(column['Name'])):
				formatteddata = datetime.strptime(datefield, '%Y-%m-%dT%H:%M:%S')
			else:
				formatteddata = result.get(column['Name'])
			resultoutput.append(formatteddata)
		wb[targetsheet].append(resultoutput)
	for sheet in wb:
		for column in columns:
			if column['Type'] == 'Date':
				cells = sheet[get_column_letter(column['Column'])]
				for cell in cells:
					if cell.row > 1:
						cell.style = 'date'
			elif column['Type'] == 'DateTime':
				cells = sheet[get_column_letter(column['Column'])]
				for cell in cells:
					if cell.row > 1:
						cell.style = 'datetime'
	try:
		wb.save(filename = args.workbook)
	except PermissionError:
		return(False)
	return(True)
		
def clear_meijer_workbook(wb):
	for ws in wb:
		ws.delete_rows(2,ws.max_row)

wb = load_meijer_workbook(args.workbook)
if not wb:
	print(f'Error loading workbook {args.workbook}')
	exit()
	
results = load_meijer_workbook_data(wb)

checkrange = list(range(int(args.startnum), int(args.endnum)+1))

for result in results:
	if (result['meijerOfferId'] in checkrange) and (args.recheck == False):
		checkrange.remove(result['meijerOfferId'])

if not checkrange:
	print('Nothing to do, exiting...')
	exit()


if not init_meijer_connection_user(args.username, args.password):
	exit()
if len(args.startnum) != 10:
	print('Wrong startnum length')
	exit()
if len(args.endnum) != 10:
	print('Wrong endnum length')
	exit()
if int(args.startnum) > int(args.endnum):
	print('startnum must be less than endnum')
	exit()

pbar = tqdm(checkrange)
for offerid in pbar:
	pbar.set_description('Scanning')
	if offerid%500 == 0:
		pbar.set_description('Processing results and saving status')
		if not process_coupons(wb):
			pbar.set_description('Error saving file...Do you have it open?')
	result = clip_meijer_coupon(offerid)
	if (not result == True) and (not result == 'Offer is already clipped by this shopper.'):
		results.append({'meijerOfferId': offerid, 'status': result, 'ScanDate': datetime.now().strftime('%Y-%m-%dT%H:%M:%S')})
print('Scanning complete, processing results')

for offertype in ['Earned', 'InProgress', 'Available']:
	print(f'Fetching {offertype} rewards')
	offers = get_meijer_offers(offertype)
	if offers:
		for offer in sorted(offers, key = lambda i: i['meijerOfferId']):
			if offer.get("meijerOfferId") in checkrange:
				results.append({'meijerOfferId': offer.get("meijerOfferId"), 'status': f'{offertype} Reward', 'description': f'{offer.get("title").strip()} {strip_tags(offer.get("description")).strip()}', 'expirationDate': offer.get("expirationDate"), 'ScanDate': datetime.now().strftime('%Y-%m-%dT%H:%M:%S')})

print('Saving...')
while not process_coupons(wb):
	input("Error saving file, press Enter to retry")
